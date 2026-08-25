import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def obter_caminho_desktop() -> Path:
    user_profile = os.environ.get("USERPROFILE")
    if user_profile:
        caminho_desktop = Path(user_profile) / "Desktop"
        if caminho_desktop.exists():
            return caminho_desktop
        caminho_at = Path(user_profile) / "Área de Trabalho"
        if caminho_at.exists():
            return caminho_at
        return Path(user_profile) / "Desktop"

    return Path.home() / "Desktop"


def gerar_relatorio_formatacao(
    registros_salvos: list[dict], produto_selecionado: str = ""
) -> Path:
    if not registros_salvos:
        return None

    agora = datetime.now()
    pasta_dia = agora.strftime("%d-%m-%Y")
    hora_formatada = agora.strftime("%Hh%M")

    nome_prod = f" {produto_selecionado.strip()}" if produto_selecionado else ""
    nome_arquivo = f"FORMATAÇÃO{nome_prod} {pasta_dia} {hora_formatada}.xlsx"

    desktop_path = obter_caminho_desktop()
    caminho_pasta_dia = desktop_path / "FORMATAÇÃO" / pasta_dia
    caminho_pasta_dia.mkdir(parents=True, exist_ok=True)

    caminho_excel = caminho_pasta_dia / nome_arquivo

    # Busca segura aceitando 'NOME ', 'NOME', 'nomeCliente', 'CPF', 'cpfCnpjCliente', etc.
    dados_simplificados = []
    for reg in registros_salvos:
        nome = (
            reg.get("NOME ")
            or reg.get("NOME")
            or reg.get("nomeCliente")
            or reg.get("Nome_Cliente")
            or ""
        )
        cpf = reg.get("CPF") or reg.get("cpfCnpjCliente") or reg.get("CPF_CNPJ") or ""
        dados_simplificados.append({"NOME": str(nome).upper(), "CPF": str(cpf)})

    wb = openpyxl.Workbook()

    # ABA 1: DADOS
    ws_dados = wb.active
    ws_dados.title = "DADOS"

    # Estilos do cabeçalho
    fill_cabecalho = PatternFill(
        start_color="002060", end_color="002060", fill_type="solid"
    )
    font_cabecalho = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_dados = Font(name="Calibri", size=11)
    alinhamento_central = Alignment(horizontal="center", vertical="center")

    borda_fina = Side(style="thin", color="000000")
    estilo_borda = Border(
        left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina
    )

    ws_dados.append(["NOME", "CPF"])

    # Formata o cabeçalho
    for col_num in range(1, 3):
        cell = ws_dados.cell(row=1, column=col_num)
        cell.fill = fill_cabecalho
        cell.font = font_cabecalho
        cell.alignment = alinhamento_central
        cell.border = estilo_borda

    # Preenche as linhas de dados com estilos
    for item in dados_simplificados:
        ws_dados.append([item["NOME"], item["CPF"]])
        row_idx = ws_dados.max_row
        for col_num in range(1, 3):
            cell = ws_dados.cell(row=row_idx, column=col_num)
            cell.font = font_dados
            cell.alignment = alinhamento_central
            cell.border = estilo_borda

    ws_dados.column_dimensions["A"].width = 40
    ws_dados.column_dimensions["B"].width = 25

    # ABA 2: PRINT
    ws_print = wb.create_sheet(title="PRINT")
    ws_print.views.sheetView[0].showGridLines = True

    wb.save(caminho_excel)

    return caminho_excel
