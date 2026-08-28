import os
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

COLUNAS_DE_DATA = ["CHEGADA DO LEAD", "DATA RECOLHE", "DATA ENTREGA"]
FORMATO_DATA_EXCEL = "YYYY-MM-DD"
FORMATO_DATA_HORA_EXECUCAO = "DD/MM/YYYY HH:MM:SS"

# --- NUMERAÇÃO INICIAL DE CADA PRODUTO SELECIONADO ---
NUMERACAO_INICIAL_POR_PRODUTO = {
    # --- PRODUTOS INRI ---
    "AUTOMOVEL - INRI": 7252,
    "BIKE - INRI": 10885,
    "CELULAR - INRI": 126085,
    "MOTO - INRI": 8126,
    "ODONTO - INRI": 211,
    "PET - INRI": 8797,
    "PORTATEIS - INRI": 8124,
    "RESIDENCIAL - INRI": 8737,
    "SAUDE - INRI": 4061,
    "VIAGEM - INRI": 56466,
    "VIDA - INRI": 1140,
    # --- PRODUTOS ICX ---
    "AUTO BONIF - ICX": 609,
    "AUTOMOVEL - ICX": 1906,
    "CELULAR - ICX": 5296,
    "CONSORCIO - ICX": 491,
    "MOTO - ICX": 10209,
    "RESIDENCIAL - ICX": 1911,
    "VIAGEM - ICX": 1456,
}

COLUNAS_MATRIZ = [
    "Nº",
    "NOME ",
    "TELEFONE",
    "CPF",
    "E-MAIL",
    "CHEGADA DO LEAD",
    "DATA RECOLHE",
    "RESPONSAVEL",
    "PRODUTO",
    "EMPRESA",
    "MODALIDADE",
    "DATA ENTREGA",
    "PROPENSAO",
    "MODELO",
    "ANO",
    "PLACA",
    "CEP",
    "UF",
    "OBSERVAÇÕES",
    "GDO",
    "idDynamics",
    "DATA/HORA EXECUÇÃO",
]


def remover_acentos(texto: str) -> str:
    """Remove acentos e caracteres especiais mantendo apenas letras normais."""
    if not texto:
        return ""
    texto_str = str(texto)
    nfkd = unicodedata.normalize("NFKD", texto_str)
    return "".join([c for c in nfkd if not unicodedata.combining(c)])


def normalizar_chave_produto(texto: str) -> str:
    """Normaliza o nome do produto para evitar falhas de digitação ou divergência de espaços/hífens."""
    if not texto:
        return ""
    txt = remover_acentos(str(texto)).upper()
    txt = re.sub(r"[^A-Z0-9]", "", txt)
    return txt


def identificar_produto_linha(registro: dict) -> str:
    """
    Identifica com precisão o produto da linha com base no campo PRODUTO e EMPRESA.
    Exemplo: Se o produto for 'MOTO' e a Empresa 'ICX', resolve para 'MOTO - ICX'.
    """
    prod_raw = str(registro.get("PRODUTO") or "").strip().upper()
    empresa_raw = str(registro.get("EMPRESA") or "").strip().upper()

    if not prod_raw:
        return "OUTROS"

    # Se a string do produto já possui o sufixo da empresa
    if "ICX" in prod_raw or "INRI" in prod_raw:
        return prod_raw

    # Se o produto veio sem a indicação da empresa, anexa a partir da coluna EMPRESA
    if "ICX" in empresa_raw:
        return f"{prod_raw} - ICX"
    elif "INRI" in empresa_raw:
        return f"{prod_raw} - INRI"

    return prod_raw


def obter_proxima_numeracao(
    df_existente: pd.DataFrame, produto_selecionado: str
) -> int:
    """
    Localiza a última numeração usada na planilha para o PRODUTO SELECIONADO.
    Se não houver registros prévios desse produto na planilha, inicia do dicionário.
    """
    prod_norm = normalizar_chave_produto(produto_selecionado)
    # 1. Pega a numeração padrão inicial configurada para o produto
    num_atual = 1
    for k, v in NUMERACAO_INICIAL_POR_PRODUTO.items():
        if normalizar_chave_produto(k) == prod_norm:
            num_atual = v
            break
    # 2. Procura no arquivo Excel a maior numeração gravada especificamente para esse produto.
    # Reconstrói a chave PRODUTO + EMPRESA de cada linha (ex: "AUTOMOVEL" + "INRI" ->
    # "AUTOMOVEL - INRI"), já que na planilha esses dados ficam em colunas separadas.
    if (
        not df_existente.empty
        and "PRODUTO" in df_existente.columns
        and "Nº" in df_existente.columns
    ):
        chave_linha = df_existente.apply(
            lambda row: normalizar_chave_produto(
                identificar_produto_linha(row.to_dict())
            ),
            axis=1,
        )
        df_produto_especifico = df_existente[chave_linha == prod_norm]
        if not df_produto_especifico.empty:
            numeros_limpos = (
                df_produto_especifico["Nº"]
                .astype(str)
                .str.replace(r"\.0$", "", regex=True)
                .str.strip()
            )
            numeros_validos = (
                pd.to_numeric(numeros_limpos, errors="coerce").dropna().astype(int)
            )
            if not numeros_validos.empty:
                num_atual = max(num_atual, numeros_validos.max())
    return num_atual


def salvar_registros_na_matriz(
    caminho_arquivo: Path, novos_registros: list[dict], produto_selecionado: str
) -> tuple[int, list[dict]]:

    agora = datetime.now()
    data_hora_execucao = agora.strftime("%Y-%m-%d %H:%M:%S")

    if caminho_arquivo.exists():
        try:
            df_existente = pd.read_excel(caminho_arquivo, dtype=str)
            col_id = (
                "idDynamics"
                if "idDynamics" in df_existente.columns
                else "idOportunidadeDynamics"
            )
            ids_cadastrados = (
                set(df_existente[col_id].dropna().astype(str).tolist())
                if col_id in df_existente.columns
                else set()
            )

            if (
                "DATA/HORA EXECUÇÃO" in df_existente.columns
                and not df_existente["DATA/HORA EXECUÇÃO"].dropna().empty
            ):
                ultimas_datas = pd.to_datetime(
                    df_existente["DATA/HORA EXECUÇÃO"], errors="coerce"
                ).dropna()
                if not ultimas_datas.empty:
                    ultima_exec = ultimas_datas.iloc[-1]
                    if (agora - ultima_exec) < timedelta(seconds=10):
                        data_hora_execucao = ultima_exec.strftime("%Y-%m-%d %H:%M:%S")
                        eh_mesmo_lote = True
                    else:
                        eh_mesmo_lote = False
                else:
                    eh_mesmo_lote = False
            else:
                eh_mesmo_lote = False

        except Exception:
            df_existente = pd.DataFrame(columns=COLUNAS_MATRIZ)
            ids_cadastrados = set()
            eh_mesmo_lote = False
    else:
        df_existente = pd.DataFrame(columns=COLUNAS_MATRIZ)
        ids_cadastrados = set()
        eh_mesmo_lote = False

    registros_filtrados = []
    duplicados_detectados = []

    # 1. Filtra registros duplicados
    for reg in novos_registros:
        id_dyn = str(reg.get("idDynamics") or reg.get("idOportunidadeDynamics") or "")
        if id_dyn in ids_cadastrados:
            duplicados_detectados.append({"id": id_dyn, "nome": reg.get("NOME ", "")})
        else:
            if reg.get("NOME "):
                reg["NOME "] = remover_acentos(reg["NOME "])
            elif reg.get("NOME"):
                reg["NOME"] = remover_acentos(reg["NOME"])

            reg["DATA/HORA EXECUÇÃO"] = data_hora_execucao
            registros_filtrados.append(reg)
            ids_cadastrados.add(id_dyn)

    if not registros_filtrados:
        return 0, duplicados_detectados

    df_novos = pd.DataFrame(registros_filtrados)

    # 2. Ordena pela data de chegada do lead antes da atribuição das numerações
    if "CHEGADA DO LEAD" in df_novos.columns:
        df_novos["CHEGADA DO LEAD"] = pd.to_datetime(
            df_novos["CHEGADA DO LEAD"], errors="coerce"
        )
        df_novos = df_novos.sort_values(
            by="CHEGADA DO LEAD", ascending=True, na_position="first"
        ).reset_index(drop=True)

    # 3. Gera a numeração sequencial isolada para cada produto selecionado
    controle_numeracao = {}
    numeros_gerados = []

    for _, row in df_novos.iterrows():
        dict_row = row.to_dict()
        prod_selecionado = identificar_produto_linha(dict_row)
        prod_norm = normalizar_chave_produto(prod_selecionado)

        # Se é a primeira ocorrência deste produto no lote atual, busca a última numeração na planilha
        if prod_norm not in controle_numeracao:
            ultimo_num = obter_proxima_numeracao(df_existente, produto_selecionado)
    numeros_gerados = []
    for _ in range(len(df_novos)):
        ultimo_num += 1
        numeros_gerados.append(ultimo_num)
    df_novos["Nº"] = numeros_gerados

    df_novos["Nº"] = numeros_gerados

    # 4. Salva os registros concatenando ao arquivo Excel
    df_final = pd.concat([df_existente, df_novos], ignore_index=True)

    for col in COLUNAS_MATRIZ:
        if col not in df_final.columns:
            df_final[col] = ""
    df_final = df_final[COLUNAS_MATRIZ]

    col_nome_matriz = "NOME " if "NOME " in df_final.columns else "NOME"
    if col_nome_matriz in df_final.columns:
        df_final[col_nome_matriz] = df_final[col_nome_matriz].apply(remover_acentos)

    for col_data in COLUNAS_DE_DATA:
        if col_data in df_final.columns:
            df_final[col_data] = pd.to_datetime(df_final[col_data], errors="coerce")

    if "DATA/HORA EXECUÇÃO" in df_final.columns:
        df_final["DATA/HORA EXECUÇÃO"] = pd.to_datetime(
            df_final["DATA/HORA EXECUÇÃO"], errors="coerce"
        )

    caminho_arquivo.parent.mkdir(parents=True, exist_ok=True)

    try:
        df_final.to_excel(caminho_arquivo, index=False)
    except PermissionError:
        raise PermissionError(
            f"Feche o arquivo '{caminho_arquivo.name}' no Excel antes de salvar!"
        )

    # 5. Formatação visual do Excel
    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb.active

    max_row = ws.max_row
    max_col_letter = get_column_letter(ws.max_column)
    ref_tabela = f"A1:{max_col_letter}{max_row}"

    tab = Table(displayName="TabelaMatrizLeads", ref=ref_tabela)
    style_info = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style_info

    ws._tables.clear()
    ws.add_table(tab)

    preenchimento_cabecalho = PatternFill(
        start_color="002060", end_color="002060", fill_type="solid"
    )
    fonte_cabecalho = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    borda_preta = Side(style="thin", color="000000")
    estilo_borda = Border(
        left=borda_preta, right=borda_preta, top=borda_preta, bottom=borda_preta
    )

    alinhamento_central = Alignment(horizontal="center", vertical="center")
    fill_sem_cor = PatternFill(fill_type=None)

    if not eh_mesmo_lote:
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = alinhamento_central
                cell.border = estilo_borda
                cell.fill = fill_sem_cor
    else:
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = alinhamento_central
                cell.border = estilo_borda

    for cell in ws[1]:
        cell.fill = preenchimento_cabecalho
        cell.font = fonte_cabecalho

    cabecalhos = [cell.value for cell in ws[1]]

    if "DATA/HORA EXECUÇÃO" in cabecalhos:
        col_idx = cabecalhos.index("DATA/HORA EXECUÇÃO") + 1
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(
                row=row_idx, column=col_idx
            ).number_format = FORMATO_DATA_HORA_EXECUCAO

    for col_data in COLUNAS_DE_DATA:
        if col_data in cabecalhos:
            col_idx = cabecalhos.index(col_data) + 1
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = FORMATO_DATA_EXCEL

    FORMATO_CPF = '000000000"-"00'
    FORMATO_CNPJ = '00"."000"."000"/"0000"-"00'
    FORMATO_TEL_11 = '"("00") "00000"-"0000'
    FORMATO_TEL_10 = '"("00") "0000"-"0000'

    idx_cpf = cabecalhos.index("CPF") + 1 if "CPF" in cabecalhos else None
    idx_tel = cabecalhos.index("TELEFONE") + 1 if "TELEFONE" in cabecalhos else None

    for row_idx in range(2, ws.max_row + 1):
        if idx_cpf:
            cell_cpf = ws.cell(row=row_idx, column=idx_cpf)
            val_doc = re.sub(r"\D", "", str(cell_cpf.value or ""))
            if val_doc:
                cell_cpf.value = int(val_doc)
                cell_cpf.number_format = (
                    FORMATO_CNPJ if len(val_doc) == 14 else FORMATO_CPF
                )

        if idx_tel:
            cell_tel = ws.cell(row=row_idx, column=idx_tel)
            val_tel = re.sub(r"\D", "", str(cell_tel.value or ""))
            if val_tel:
                cell_tel.value = int(val_tel)
                if len(val_tel) == 11:
                    cell_tel.number_format = FORMATO_TEL_11
                elif len(val_tel) == 10:
                    cell_tel.number_format = FORMATO_TEL_10

    preenchimento_amarelo = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    idx_execucao = (
        cabecalhos.index("DATA/HORA EXECUÇÃO") + 1
        if "DATA/HORA EXECUÇÃO" in cabecalhos
        else None
    )

    if idx_execucao:
        dt_exec_obj = datetime.strptime(data_hora_execucao, "%Y-%m-%d %H:%M:%S")
        for row_idx in range(2, ws.max_row + 1):
            val_exec = ws.cell(row=row_idx, column=idx_execucao).value

            if isinstance(val_exec, str):
                try:
                    val_exec = datetime.strptime(val_exec, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    pass

            if isinstance(val_exec, datetime) and val_exec == dt_exec_obj:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = preenchimento_amarelo
                break

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if cell.number_format in [FORMATO_TEL_11, FORMATO_TEL_10]:
                val_str = "(00) 00000-0000"
            elif cell.number_format == FORMATO_CPF:
                val_str = "000000000-00"
            elif cell.number_format == FORMATO_CNPJ:
                val_str = "00.000.000/0000-00"
            elif cell.number_format == FORMATO_DATA_HORA_EXECUCAO:
                val_str = "DD/MM/YYYY HH:MM:SS"

            max_len = max(max_len, len(val_str))

        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(caminho_arquivo)

    try:
        os.startfile(caminho_arquivo)
    except Exception:
        pass

    return len(df_novos), duplicados_detectados
